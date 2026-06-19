#!/usr/bin/env python3
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import calendar

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from netCDF4 import Dataset, num2date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from scipy import stats as scipy_stats

from compare_sd_mwr_scatter import (
    compute_mwr_modwt,
    load_mwr_mat,
    load_sd_annual,
    ut_to_lt,
)


@dataclass(frozen=True)
class CaseConfig:
    name: str
    year: int
    sd_code: str
    mwr_label: str
    mwr_mat: Path
    jawara_sd_nc: Path
    jawara_mwr_nc: Path
    climit: tuple[float, float]


CASES = [
    CaseConfig(
        name="han_and",
        year=2008,
        sd_code="han",
        mwr_label="AND",
        mwr_mat=Path("~/data/meteor_winds/mat/And_2008.mat").expanduser(),
        jawara_sd_nc=Path("~/data/meteor_winds/jawara/han_2008_jawara_hourly_uvz_2x2.nc").expanduser(),
        jawara_mwr_nc=Path("~/data/meteor_winds/jawara/han_2008_mwr_jawara_hourly_uvz_2x2.nc").expanduser(),
        climit=(-70.0, 70.0),
    ),
    CaseConfig(
        name="mcm_mcm",
        year=2019,
        sd_code="mcm",
        mwr_label="MCM",
        mwr_mat=Path("~/data/meteor_winds/mat/McMurdo_2019.mat").expanduser(),
        jawara_sd_nc=Path("~/data/meteor_winds/jawara/mcm_2019_jawara_hourly_uvz_2x2.nc").expanduser(),
        jawara_mwr_nc=Path("~/data/meteor_winds/jawara/mcm_2019_jawara_hourly_uvz_2x2.nc").expanduser(),
        climit=(-50.0, 50.0),
    ),
]

LOW_COVERAGE_MIN_HOURS = 1


def moving_nanmedian(arr: np.ndarray, window: int) -> np.ndarray:
    out = np.full_like(arr, np.nan, dtype=float)
    half = window // 2
    for idx in range(arr.shape[1]):
        lo = max(0, idx - half)
        hi = min(arr.shape[1], idx + half + 1)
        window_slice = arr[:, lo:hi]
        for row in range(arr.shape[0]):
            if np.any(np.isfinite(window_slice[row, :])):
                out[row, idx] = np.nanmedian(window_slice[row, :])
    return out


def month_ticks(year: int) -> tuple[np.ndarray, list[str]]:
    starts = []
    labels = []
    doy = 1
    for month in range(1, 13):
        starts.append(doy)
        labels.append(calendar.month_abbr[month])
        doy += calendar.monthrange(year, month)[1]
    return np.asarray(starts, dtype=float), labels


def paper_rgb_colormap() -> np.ndarray:
    base = np.array(
        [
            [94, 79, 162],
            [50, 136, 189],
            [102, 194, 165],
            [171, 221, 164],
            [230, 245, 152],
            [255, 255, 191],
            [254, 224, 139],
            [253, 174, 97],
            [244, 109, 67],
            [213, 62, 79],
            [158, 1, 66],
        ],
        dtype=float,
    ) / 255.0
    xi = np.linspace(0, len(base) - 1, len(base) * 10)
    xp = np.arange(len(base), dtype=float)
    cmap = np.column_stack([np.interp(xi, xp, base[:, idx]) for idx in range(3)])
    return cmap


def geopotential_to_geometric_height(height_km: np.ndarray, earth_radius_km: float = 6371.0) -> np.ndarray:
    height_km = np.asarray(height_km, dtype=float)
    out = np.full_like(height_km, np.nan, dtype=float)
    valid = np.isfinite(height_km) & (height_km < earth_radius_km)
    out[valid] = earth_radius_km * height_km[valid] / (earth_radius_km - height_km[valid])
    return out


def apply_day_mask(arr: np.ndarray, good_days: np.ndarray) -> np.ndarray:
    out = np.array(arr, dtype=float, copy=True)
    out[:, ~good_days] = np.nan
    return out


def bilinear_weights(
    lat_grid: np.ndarray,
    lon_grid: np.ndarray,
    target_lat: float,
    target_lon: float,
) -> np.ndarray:
    target_lon = target_lon % 360.0
    lat0 = float(lat_grid[0])
    lat1 = float(lat_grid[1])
    lon0 = float(lon_grid[0])
    lon1 = float(lon_grid[1])
    fy = 0.0 if lat1 == lat0 else (target_lat - lat0) / (lat1 - lat0)
    fx = 0.0 if lon1 == lon0 else (target_lon - lon0) / (lon1 - lon0)
    fy = float(np.clip(fy, 0.0, 1.0))
    fx = float(np.clip(fx, 0.0, 1.0))
    return np.array(
        [
            [(1.0 - fy) * (1.0 - fx), (1.0 - fy) * fx],
            [fy * (1.0 - fx), fy * fx],
        ],
        dtype=float,
    )


def load_mwr_full_year(cfg: CaseConfig) -> dict[str, np.ndarray]:
    raw = load_mwr_mat(cfg.mwr_mat)
    n_days = 366 if calendar.isleap(cfg.year) else 365
    u = np.full((raw["u"].shape[0], raw["u"].shape[1], n_days), np.nan, dtype=float)
    v = np.full_like(u, np.nan)
    for src_idx, day in enumerate(raw["day_doy"]):
        if not np.isfinite(day):
            continue
        day_index = int(day) - 1
        if 0 <= day_index < n_days:
            u[:, :, day_index] = raw["u"][:, :, src_idx]
            v[:, :, day_index] = raw["v"][:, :, src_idx]
    day_doy = np.arange(1, n_days + 1, dtype=float)
    return {
        "u": u,
        "v": v,
        "alt": raw["alt"],
        "hour": raw["hour"],
        "day_doy": day_doy,
        "lat": float(raw["lat"]),
        "lon": float(raw["lon"]),
    }


def load_jawara_weighted(jawara_nc: Path, year: int, sd: dict[str, np.ndarray]) -> dict[str, np.ndarray]:
    with Dataset(jawara_nc, "r") as ds:
        time_var = ds.variables["time"]
        times = num2date(time_var[:], units=time_var.units)
        z = np.asarray(ds.variables["z"][:], dtype=float) / 1000.0
        u = np.asarray(ds.variables["u"][:], dtype=float)
        v = np.asarray(ds.variables["v"][:], dtype=float)
        lat_grid = np.asarray(ds.variables["latitude"][:], dtype=float)
        lon_grid = np.asarray(ds.variables["longitude"][:], dtype=float)
        target_lat = float(ds.getncattr("target_latitude"))
        target_lon = float(ds.getncattr("target_longitude"))

    weights_xy = bilinear_weights(lat_grid, lon_grid, target_lat, target_lon)
    z_site = np.tensordot(z, weights_xy, axes=([2, 3], [0, 1]))
    u_site = np.tensordot(u, weights_xy, axes=([2, 3], [0, 1]))
    v_site = np.tensordot(v, weights_xy, axes=([2, 3], [0, 1]))
    z_site = geopotential_to_geometric_height(z_site)

    n_days = 366 if calendar.isleap(year) else 365
    u_modwt = np.full((24, n_days), np.nan, dtype=float)
    v_modwt = np.full((24, n_days), np.nan, dtype=float)
    u_sigma = np.full((24, n_days), np.nan, dtype=float)
    v_sigma = np.full((24, n_days), np.nan, dtype=float)

    for ti, dt in enumerate(times):
        day_index = int(dt.timetuple().tm_yday) - 1
        hour_index = int(dt.hour)
        peak = sd["peak"][hour_index, day_index]
        fwhm = sd["fwhm"][hour_index, day_index]
        if not np.isfinite(peak) or not np.isfinite(fwhm) or fwhm <= 0:
            continue
        sigma = fwhm / 2.0
        z_prof = z_site[ti, :]
        valid = np.isfinite(z_prof)
        if not np.any(valid):
            continue
        weights_z = np.exp(-0.5 * ((z_prof[valid] - peak) / sigma) ** 2)
        if np.nansum(weights_z) <= 0:
            continue

        u_prof = u_site[ti, valid]
        v_prof = v_site[ti, valid]
        if np.any(np.isfinite(u_prof)):
            good = np.isfinite(u_prof)
            wu = weights_z[good]
            u_mean = np.nansum(u_prof[good] * wu) / np.nansum(wu)
            u_modwt[hour_index, day_index] = u_mean
            u_sigma[hour_index, day_index] = np.sqrt(
                np.nansum(wu * (u_prof[good] - u_mean) ** 2) / np.nansum(wu)
            )
        if np.any(np.isfinite(v_prof)):
            good = np.isfinite(v_prof)
            wv = weights_z[good]
            v_mean = np.nansum(v_prof[good] * wv) / np.nansum(wv)
            v_modwt[hour_index, day_index] = v_mean
            v_sigma[hour_index, day_index] = np.sqrt(
                np.nansum(wv * (v_prof[good] - v_mean) ** 2) / np.nansum(wv)
            )

    return {
        "u_modwt": u_modwt,
        "v_modwt": v_modwt,
        "u_sigma": u_sigma,
        "v_sigma": v_sigma,
        "lat": target_lat,
        "lon": target_lon,
    }


def linfit_stats(x: np.ndarray, y: np.ndarray) -> dict[str, float]:
    mask = np.isfinite(x) & np.isfinite(y)
    if np.count_nonzero(mask) < 2:
        return {"slope": np.nan, "intercept": np.nan, "corr": np.nan, "p_value": np.nan, "n": float(np.count_nonzero(mask))}
    slope, intercept = np.polyfit(x[mask], y[mask], 1)
    corr, p_value = scipy_stats.pearsonr(x[mask], y[mask])
    if p_value == 0.0:
        p_value = np.nextafter(0.0, 1.0)
    return {
        "slope": float(slope),
        "intercept": float(intercept),
        "corr": float(corr),
        "p_value": float(p_value),
        "n": float(mask.sum()),
    }


def plot_contours(
    out_path: Path,
    cfg: CaseConfig,
    mwr_u: np.ndarray,
    mwr_v: np.ndarray,
    sd_u: np.ndarray,
    sd_v: np.ndarray,
    jaw_u: np.ndarray,
    jaw_v: np.ndarray,
    mwr_lat: float,
    mwr_lon: float,
    sd_lat: float,
    sd_lon: float,
    good_days: np.ndarray,
) -> None:
    fig, axes = plt.subplots(2, 3, figsize=(12.5, 6.7), constrained_layout=True, sharey=True)
    levels = np.linspace(cfg.climit[0], cfg.climit[1], 25)
    line_levels = np.arange(cfg.climit[0], cfg.climit[1] + 1, 20.0)
    cmap = plt.matplotlib.colors.ListedColormap(paper_rgb_colormap())
    cmap.set_bad("#d9d9d9")
    x_mwr = np.arange(1, mwr_u.shape[1] + 1)
    x_sd = np.arange(1, sd_u.shape[1] + 1)
    x_jaw = np.arange(1, jaw_u.shape[1] + 1)
    y = np.arange(24)
    xticks, xticklabels = month_ticks(cfg.year)

    panels = [
        (axes[0, 0], x_mwr, mwr_u, f"{cfg.mwr_label} Meteor Radar ({mwr_lat:.1f}, {mwr_lon:.1f})", "Zonal\nLST (hr)", False),
        (axes[0, 1], x_sd, sd_u, f"{cfg.sd_code.upper()} SuperDARN ({sd_lat:.1f}, {sd_lon:.1f})", None, True),
        (axes[0, 2], x_jaw, jaw_u, f"JAWARA @ {cfg.sd_code.upper()}", None, False),
        (axes[1, 0], x_mwr, mwr_v, None, r"$\bf{Meridional}$" + "\nLST (hr)", False),
        (axes[1, 1], x_sd, sd_v, None, None, True),
        (axes[1, 2], x_jaw, jaw_v, None, None, False),
    ]
    panels[0] = (panels[0][0], panels[0][1], panels[0][2], panels[0][3], r"$\bf{Zonal}$" + "\nLST (hr)", panels[0][5])

    mesh = None
    for idx, (ax, x, data, title, ylabel, mask_sd_only) in enumerate(panels):
        if mask_sd_only:
            display = apply_day_mask(data, good_days)
        else:
            display = data
        masked = np.ma.masked_invalid(display)
        mesh = ax.contourf(x, y, masked, levels=levels, cmap=cmap, extend="both")
        cs = ax.contour(x, y, masked, levels=line_levels, colors="0.25", linewidths=0.45, alpha=0.65)
        ax.clabel(cs, fmt="%d", fontsize=7, inline=True, inline_spacing=2)
        ax.set_ylim(0, 23)
        ax.set_yticks([0, 6, 12, 18, 23])
        ax.set_xticks(xticks)
        ax.set_xticklabels(xticklabels, fontsize=9)
        if title:
            ax.set_title(title, fontsize=11)
        if ylabel:
            ax.set_ylabel(ylabel)
        if idx < 3:
            ax.set_xticklabels([])
        ax.grid(True, alpha=0.12, linewidth=0.4)

        if mask_sd_only:
            for bad_idx in np.where(~good_days)[0]:
                ax.axvspan(bad_idx + 0.5, bad_idx + 1.5, color="#d9d9d9", alpha=0.35, linewidth=0)

    cbar = fig.colorbar(mesh, ax=axes, shrink=0.96, pad=0.02)
    cbar.set_label("Wind (m/s)")
    fig.savefig(out_path, dpi=220, bbox_inches="tight")
    plt.close(fig)


def plot_scatter(
    out_path: Path,
    cfg: CaseConfig,
    lt_mwr_u: np.ndarray,
    lt_mwr_v: np.ndarray,
    lt_sd_u: np.ndarray,
    lt_sd_v: np.ndarray,
    lt_jaw_sd_u: np.ndarray,
    lt_jaw_sd_v: np.ndarray,
    lt_jaw_mwr_u: np.ndarray,
    lt_jaw_mwr_v: np.ndarray,
) -> list[dict[str, float | str]]:
    fig, axes = plt.subplots(2, 3, figsize=(13.2, 8.2), constrained_layout=True)
    panels = [
        ("Meteor Radar vs SuperDARN", "Zonal", lt_mwr_u, lt_sd_u, axes[0, 0], "mwr", "Meteor radar (m/s)", "SuperDARN (m/s)"),
        ("JAWARA vs SuperDARN", "Zonal", lt_jaw_sd_u, lt_sd_u, axes[0, 1], "jawara", "JAWARA (m/s)", "SuperDARN (m/s)"),
        ("Meteor Radar vs JAWARA", "Zonal", lt_mwr_u, lt_jaw_mwr_u, axes[0, 2], "mwr_vs_jawara", "Meteor radar (m/s)", "JAWARA (m/s)"),
        ("Meteor Radar vs SuperDARN", "Meridional", lt_mwr_v, lt_sd_v, axes[1, 0], "mwr", "Meteor radar (m/s)", "SuperDARN (m/s)"),
        ("JAWARA vs SuperDARN", "Meridional", lt_jaw_sd_v, lt_sd_v, axes[1, 1], "jawara", "JAWARA (m/s)", "SuperDARN (m/s)"),
        ("Meteor Radar vs JAWARA", "Meridional", lt_mwr_v, lt_jaw_mwr_v, axes[1, 2], "mwr_vs_jawara", "Meteor radar (m/s)", "JAWARA (m/s)"),
    ]

    rows: list[dict[str, float | str]] = []
    for heading, component, xgrid, ygrid, ax, comparison, xlabel, ylabel in panels:
        x = xgrid.ravel()
        y = ygrid.ravel()
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        stats = linfit_stats(x, y)
        ax.scatter(x, y, s=6, alpha=0.18, color="#1f77b4", edgecolors="none")
        if x.size:
            lo = min(np.nanmin(x), np.nanmin(y))
            hi = max(np.nanmax(x), np.nanmax(y))
            pad = 0.05 * (hi - lo) if hi > lo else 1.0
            xx = np.array([lo - pad, hi + pad])
            ax.plot(xx, xx, "--", color="0.25", linewidth=1.0)
            ax.plot(xx, stats["slope"] * xx + stats["intercept"], color="#d62728", linewidth=1.2)
            ax.set_xlim(xx[0], xx[1])
            ax.set_ylim(xx[0], xx[1])
        ax.set_title(f"{heading} {component}", fontsize=11)
        ax.set_xlabel(xlabel)
        ax.set_ylabel(ylabel)
        ax.grid(True, alpha=0.2)
        ax.text(
            0.03,
            0.97,
            f"slope={stats['slope']:.2f}\nintercept={stats['intercept']:.1f}\nr={stats['corr']:.2f}\np={stats['p_value']:.2g}\nn={stats['n']:.0f}",
            transform=ax.transAxes,
            va="top",
            ha="left",
            fontsize=9,
            bbox={"facecolor": "white", "edgecolor": "0.8", "alpha": 0.85, "pad": 3},
        )
        rows.append(
            {
                "case": cfg.name,
                "comparison": comparison,
                "component": component.lower(),
                **stats,
            }
        )

    fig.suptitle(f"{cfg.year}: Meteor radar, SuperDARN, and JAWARA comparisons for {cfg.name}", fontsize=13)
    fig.savefig(out_path, dpi=220, bbox_inches="tight")
    plt.close(fig)
    return rows


def save_stats_table(out_path: Path, rows: list[dict[str, float | str]]) -> None:
    keys = ["case", "comparison", "component", "slope", "intercept", "corr", "p_value", "n"]
    with out_path.open("w", encoding="utf-8") as handle:
        handle.write("\t".join(keys) + "\n")
        for row in rows:
            vals = []
            for key in keys:
                value = row[key]
                if isinstance(value, str):
                    vals.append(value)
                elif key == "n":
                    vals.append(f"{value:.0f}")
                elif key == "p_value":
                    vals.append(f"{value:.3e}")
                else:
                    vals.append(f"{value:.3f}")
            handle.write("\t".join(vals) + "\n")


def save_case_summary(out_path: Path, rows: list[dict[str, float | str]]) -> None:
    cases = []
    for row in rows:
        case = str(row["case"])
        if case not in cases:
            cases.append(case)

    with out_path.open("w", encoding="utf-8") as handle:
        handle.write(
            "case\tstat\tmwr_vs_sd_zonal\tmwr_vs_sd_meridional\tjawara_vs_sd_zonal\tjawara_vs_sd_meridional\tmwr_vs_jawara_zonal\tmwr_vs_jawara_meridional\n"
        )
        for case in cases:
            case_rows = {
                (str(row["comparison"]), str(row["component"])): row
                for row in rows
                if str(row["case"]) == case
            }
            for stat in ["slope", "intercept", "corr", "p_value", "n"]:
                vals = [case, stat]
                for comp, direction in [
                    ("mwr", "zonal"),
                    ("mwr", "meridional"),
                    ("jawara", "zonal"),
                    ("jawara", "meridional"),
                    ("mwr_vs_jawara", "zonal"),
                    ("mwr_vs_jawara", "meridional"),
                ]:
                    value = case_rows[(comp, direction)][stat]
                    if stat == "n":
                        vals.append(f"{float(value):.0f}")
                    elif stat == "p_value":
                        vals.append(f"{float(value):.3e}")
                    else:
                        vals.append(f"{float(value):.3f}")
                handle.write("\t".join(vals) + "\n")


def build_case_sheet(case: str, rows: list[dict[str, float | str]]) -> pd.DataFrame:
    case_rows = {
        (str(row["comparison"]), str(row["component"])): row
        for row in rows
        if str(row["case"]) == case
    }
    index = ["Slope", "Y-intercept (m/s)", "Correlation", "p-value", "Samples"]
    data = {
        "MWR vs SuperDARN Zonal": [
            case_rows[("mwr", "zonal")]["slope"],
            case_rows[("mwr", "zonal")]["intercept"],
            case_rows[("mwr", "zonal")]["corr"],
            case_rows[("mwr", "zonal")]["p_value"],
            case_rows[("mwr", "zonal")]["n"],
        ],
        "MWR vs SuperDARN Meridional": [
            case_rows[("mwr", "meridional")]["slope"],
            case_rows[("mwr", "meridional")]["intercept"],
            case_rows[("mwr", "meridional")]["corr"],
            case_rows[("mwr", "meridional")]["p_value"],
            case_rows[("mwr", "meridional")]["n"],
        ],
        "JAWARA vs SuperDARN Zonal": [
            case_rows[("jawara", "zonal")]["slope"],
            case_rows[("jawara", "zonal")]["intercept"],
            case_rows[("jawara", "zonal")]["corr"],
            case_rows[("jawara", "zonal")]["p_value"],
            case_rows[("jawara", "zonal")]["n"],
        ],
        "JAWARA vs SuperDARN Meridional": [
            case_rows[("jawara", "meridional")]["slope"],
            case_rows[("jawara", "meridional")]["intercept"],
            case_rows[("jawara", "meridional")]["corr"],
            case_rows[("jawara", "meridional")]["p_value"],
            case_rows[("jawara", "meridional")]["n"],
        ],
        "MWR vs JAWARA Zonal": [
            case_rows[("mwr_vs_jawara", "zonal")]["slope"],
            case_rows[("mwr_vs_jawara", "zonal")]["intercept"],
            case_rows[("mwr_vs_jawara", "zonal")]["corr"],
            case_rows[("mwr_vs_jawara", "zonal")]["p_value"],
            case_rows[("mwr_vs_jawara", "zonal")]["n"],
        ],
        "MWR vs JAWARA Meridional": [
            case_rows[("mwr_vs_jawara", "meridional")]["slope"],
            case_rows[("mwr_vs_jawara", "meridional")]["intercept"],
            case_rows[("mwr_vs_jawara", "meridional")]["corr"],
            case_rows[("mwr_vs_jawara", "meridional")]["p_value"],
            case_rows[("mwr_vs_jawara", "meridional")]["n"],
        ],
    }
    return pd.DataFrame(data, index=index)


def save_excel_tables(out_path: Path, rows: list[dict[str, float | str]]) -> None:
    def cell_text(metric: str, value: float) -> str:
        if metric == "Slope":
            return f"{value:.2f}"
        if metric == "Y-intercept (m/s)":
            return f"{value:.1f}"
        if metric == "Correlation":
            return f"{value:.2f}"
        if metric == "Samples":
            return f"{int(round(value))}"
        if metric == "p-value":
            return "<1e-300" if value < 1e-300 else f"{value:.2e}"
        return str(value)

    def add_sheet(wb: Workbook, title: str, case: str) -> None:
        ws = wb.create_sheet(title)
        case_rows = {
            (str(row["comparison"]), str(row["component"])): row
            for row in rows
            if str(row["case"]) == case
        }

        ws.merge_cells("B1:C1")
        ws.merge_cells("D1:E1")
        ws.merge_cells("F1:G1")
        ws["B1"] = "MWR vs SuperDARN"
        ws["D1"] = "JAWARA vs SuperDARN"
        ws["F1"] = "MWR vs JAWARA"
        ws["B2"] = "Zonal"
        ws["C2"] = "Meridional"
        ws["D2"] = "Zonal"
        ws["E2"] = "Meridional"
        ws["F2"] = "Zonal"
        ws["G2"] = "Meridional"

        metrics = [
            ("Slope", "slope"),
            ("Y-intercept (m/s)", "intercept"),
            ("Correlation", "corr"),
            ("p-value", "p_value"),
            ("Samples", "n"),
        ]
        comp_order = [
            ("mwr", "zonal"),
            ("mwr", "meridional"),
            ("jawara", "zonal"),
            ("jawara", "meridional"),
            ("mwr_vs_jawara", "zonal"),
            ("mwr_vs_jawara", "meridional"),
        ]
        for row_idx, (metric_label, stat_key) in enumerate(metrics, start=3):
            ws.cell(row=row_idx, column=1, value=metric_label)
            for col_idx, comp_key in enumerate(comp_order, start=2):
                value = float(case_rows[comp_key][stat_key])
                ws.cell(row=row_idx, column=col_idx, value=cell_text(metric_label, value))

        header_font = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in ws.iter_rows(min_row=1, max_row=7, min_col=1, max_col=7):
            for cell in row:
                cell.border = border
                if cell.row <= 2:
                    cell.font = header_font
                    cell.alignment = center
                elif cell.column == 1:
                    cell.font = header_font
                else:
                    cell.alignment = center

        ws.column_dimensions["A"].width = 20
        for col in ["B", "C", "D", "E", "F", "G"]:
            ws.column_dimensions[col].width = 18
        ws.freeze_panes = "B3"

    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb, "Table5", "han_and")
    add_sheet(wb, "Table6", "mcm_mcm")

    ws = wb.create_sheet("LongForm")
    long_df = pd.DataFrame(rows)
    ws.append(list(long_df.columns))
    for row in long_df.itertuples(index=False):
        out = []
        for key, value in zip(long_df.columns, row):
            if isinstance(value, str):
                out.append(value)
            elif key == "n":
                out.append(int(round(float(value))))
            elif key == "p_value":
                out.append("<1e-300" if float(value) < 1e-300 else f"{float(value):.2e}")
            else:
                out.append(float(value))
        ws.append(out)

    wb.save(out_path)


def main() -> None:
    plt.rcParams.update(
        {
            "font.size": 10,
            "axes.titlesize": 11,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
        }
    )

    out_dir = Path("/Users/chartat1/superdarn/winds/outputs/paper_jawara")
    out_dir.mkdir(parents=True, exist_ok=True)
    all_rows: list[dict[str, float | str]] = []
    lthri = np.arange(24, dtype=float)

    for cfg in CASES:
        sd_path = Path(f"/Users/chartat1/data/superdarn/fit_nc_3_winds/annual/{cfg.year}/{cfg.sd_code}_{cfg.year}.nc")
        sd = load_sd_annual(sd_path, cfg.sd_code)
        good_days = np.sum(np.isfinite(sd["u"]), axis=0) >= LOW_COVERAGE_MIN_HOURS
        mwr = load_mwr_full_year(cfg)
        jawara_sd = load_jawara_weighted(cfg.jawara_sd_nc, cfg.year, sd)
        jawara_mwr = load_jawara_weighted(cfg.jawara_mwr_nc, cfg.year, sd)

        mwr_for_weight = {
            "u": mwr["u"],
            "v": mwr["v"],
            "alt": mwr["alt"],
            "hour": mwr["hour"],
            "day_doy": mwr["day_doy"],
            "lat": mwr["lat"],
            "lon": mwr["lon"],
        }
        mwr_u, mwr_v, _, _ = compute_mwr_modwt(mwr_for_weight, sd)

        mwr_u_med = moving_nanmedian(mwr_u, 31)
        mwr_v_med = moving_nanmedian(mwr_v, 31)
        sd_u_med = moving_nanmedian(sd["u"], 31)
        sd_v_med = moving_nanmedian(sd["v"], 31)
        jaw_sd_u_med = moving_nanmedian(jawara_sd["u_modwt"], 31)
        jaw_sd_v_med = moving_nanmedian(jawara_sd["v_modwt"], 31)
        jaw_mwr_u_med = moving_nanmedian(jawara_mwr["u_modwt"], 31)
        jaw_mwr_v_med = moving_nanmedian(jawara_mwr["v_modwt"], 31)

        lt_mwr_u = ut_to_lt(mwr_u_med, mwr["hour"], lthri, mwr["lon"])
        lt_mwr_v = ut_to_lt(mwr_v_med, mwr["hour"], lthri, mwr["lon"])
        lt_sd_u = ut_to_lt(sd_u_med, sd["hour"], lthri, sd["lon"])
        lt_sd_v = ut_to_lt(sd_v_med, sd["hour"], lthri, sd["lon"])
        lt_jaw_sd_u = ut_to_lt(jaw_sd_u_med, sd["hour"], lthri, jawara_sd["lon"])
        lt_jaw_sd_v = ut_to_lt(jaw_sd_v_med, sd["hour"], lthri, jawara_sd["lon"])
        lt_jaw_mwr_u = ut_to_lt(jaw_mwr_u_med, mwr["hour"], lthri, jawara_mwr["lon"])
        lt_jaw_mwr_v = ut_to_lt(jaw_mwr_v_med, mwr["hour"], lthri, jawara_mwr["lon"])

        contour_path = out_dir / f"{cfg.name}_sd_mwr_jawara_contours.png"
        plot_contours(
            contour_path,
            cfg,
            lt_mwr_u,
            lt_mwr_v,
            lt_sd_u,
            lt_sd_v,
            lt_jaw_sd_u,
            lt_jaw_sd_v,
            mwr["lat"],
            mwr["lon"],
            sd["lat"],
            sd["lon"],
            good_days,
        )

        scatter_path = out_dir / f"{cfg.name}_sd_mwr_jawara_scatter.png"
        lt_mwr_u_masked = apply_day_mask(lt_mwr_u, good_days)
        lt_mwr_v_masked = apply_day_mask(lt_mwr_v, good_days)
        lt_sd_u_masked = apply_day_mask(lt_sd_u, good_days)
        lt_sd_v_masked = apply_day_mask(lt_sd_v, good_days)
        lt_jaw_sd_u_masked = apply_day_mask(lt_jaw_sd_u, good_days)
        lt_jaw_sd_v_masked = apply_day_mask(lt_jaw_sd_v, good_days)
        lt_jaw_mwr_u_masked = apply_day_mask(lt_jaw_mwr_u, good_days)
        lt_jaw_mwr_v_masked = apply_day_mask(lt_jaw_mwr_v, good_days)
        all_rows.extend(
            plot_scatter(
                scatter_path,
                cfg,
                lt_mwr_u_masked,
                lt_mwr_v_masked,
                lt_sd_u_masked,
                lt_sd_v_masked,
                lt_jaw_sd_u_masked,
                lt_jaw_sd_v_masked,
                lt_jaw_mwr_u_masked,
                lt_jaw_mwr_v_masked,
            )
        )

    save_stats_table(out_dir / "paper_jawara_stats.tsv", all_rows)
    save_case_summary(out_dir / "paper_jawara_table5_table6_replacement.tsv", all_rows)
    save_excel_tables(out_dir / "paper_jawara_table5_table6_replacement.xlsx", all_rows)


if __name__ == "__main__":
    main()
